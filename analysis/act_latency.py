#!/usr/bin/python
# ------------------------------------------------
# act_latency.py
#
# Analyze an act_storage or act_index output file.
# Typical print_usage:
#    $ ./act_latency.py -l act_out.txt
# where act_out.txt is output generated by act_storage or act_index, and which
# uses defaults:
# (-h - depends on config found in act_out.txt)
# -t 3600
# -s 0
# -n 7
# -e 1
# (-x - not set)
# ------------------------------------------------


# ==========================================================
# Imports.
#

from __future__ import print_function
import getopt
import re
import sys
from datetime import date

try:
    from openpyxl.styles import NamedStyle, Font, Border, Side, colors, Color
    from openpyxl import Workbook
    from openpyxl.chart import (
        LineChart,
        Reference,
    )
    from copy import deepcopy
    have_openpyxl = True
except:
    have_openpyxl = False
    

# ==========================================================
# Compatibility.
#

if sys.version_info[0] == 3:
    long = int


# ==========================================================
# Constants.
#

BUCKET_LABELS = ("00", "01", "02", "03", "04", "05", "06", "07", "08", "09",
                 "10", "11", "12", "13", "14", "15", "16")
ALL_BUCKETS = len(BUCKET_LABELS)
BUCKET_PATTERNS = [re.compile('.*?\(' + b + ': (.*?)\).*?')
                   for b in BUCKET_LABELS]
GAP = "  "


class Args(object):
    log = None
    excel = None
    histograms = []
    slice = 3600
    start_bucket = 0
    num_buckets = 7
    every_nth = 1
    extra = False
    graph = False


class Hist(object):
    scale_label = ""
    underline = ""
    max_bucket = 0
    bucket_range = None
    display_range = None
    slice_time = 0
    first_table_row = None # first row for slice tables
    first_col = 2
    cur_row = None
    last_table_row = None

    def __init__(self, name):
        self.name = name

        self.pre_pad = ""
        self.old_total = 0
        self.old_values = [0] * Hist.max_bucket
        self.slice_total = 0
        self.slice_values = [0] * Hist.max_bucket
        self.rate = 0.0
        self.avg_rate = 0.0
        self.max_rate = 0.0
        self.overs = [0.0] * Hist.max_bucket
        self.avg_overs = [0.0] * Hist.max_bucket
        self.max_overs = [0.0] * Hist.max_bucket
        self.start_col = None


# ==========================================================
# Main.
#

def main():
    get_args()
    config = open_log_file()
    file_id = config['misc']['file_id']
    
    find_max_bucket()
    hists = [Hist(name) for name in Args.histograms]

    if Args.excel is not None:
        if not Args.excel.endswith('.xlsx'):
            Args.excel += '.xlsx'
        try:
            wb = Workbook()
            wb.save(Args.excel)
        except:
            print("can't save " + Args.excel + ": check permissions")
            print_usage()
            sys.exit(-1)
        Hist.first_table_row = excel_config(config, wb, 2, Hist.first_col) + 1
        excel_table_header(hists, wb)
    else:
        wb = None
        print_config(config)
        print_table_header(hists)

    num_slices = output_latency_slices(hists, file_id, wb)
    if Args.excel is not None:
        excel_latency_aggregates(hists, num_slices, wb)

        fr = Hist.first_table_row
        lr = Hist.last_table_row
        fc = Hist.first_col

        # box around Slice label column
        thicken(wb.active,  fr + 2,  fc,     fr + 2,  fc)
        thicken(wb.active,  fr + 3,  fc,     lr,      fc)
        thicken(wb.active,  lr + 1,  fc,     lr + 2,  fc)

        for hist in hists:
            # label box
            lc = fc + len(Hist.display_range) + Args.extra
            thicken(wb.active,  fr + 2,  fc + 1, fr + 2,  lc)
            # slices box
            thicken(wb.active,  fr + 3,  fc + 1, lr,      lc)
            # aggregate box
            thicken(wb.active,  lr + 1,  fc + 1, lr + 2,  lc)
            fc = lc + 1

        if (Args.graph):
            excel_graph(wb, hists, 1, 1)
        
        wb.save(Args.excel)
    else:
        print_latency_aggregates(hists, num_slices)


# ==========================================================
# Helper functions.
#

# ------------------------------------------------
# Get and sanity-check command line arguments.
#
def get_args():
    # Echo the command line arguments.
    print("act_latency.py " + " ".join(sys.argv[1:]))

    # Read the input arguments:
    try:
        opts, args = getopt.getopt(
            sys.argv[1:], "l:h:t:s:n:e:c:xg",
            ["log=", "histogram=", "slice=", "start_bucket=", "num_buckets=",
             "every_nth=", "extra"])
    except getopt.GetoptError as err:
        print(str(err))
        print_usage()
        sys.exit(-1)

    # Set the arguments:
    for o, a in opts:
        if o == "-l" or o == "--log":
            Args.log = a
        elif o == "-h" or o == "--histogram":
            Args.histograms.append(a)
        elif o == "-t" or o == "--slice":
            Args.slice = long(a)
        elif o == "-s" or o == "--start_bucket":
            Args.start_bucket = int(a)
        elif o == "-n" or o == "--num_buckets":
            Args.num_buckets = int(a)
        elif o == "-e" or o == "--every_nth":
            Args.every_nth = int(a)
        elif o == "-c" or o == "--excel":
            Args.excel = a
        elif o == "-x" or o == "--extra":
            Args.extra = True
        elif o == "-g" or o == "--graph":
            Args.graph = True

    # Sanity-check the arguments:
    if Args.log is None:
        print_usage()
        sys.exit(-1)

    if Args.slice < 1:
        print("slice must be more than 0")
        sys.exit(-1)

    if Args.start_bucket < 0 or Args.start_bucket >= ALL_BUCKETS:
        print("start_bucket must be non-negative and less than " + ALL_BUCKETS)
        sys.exit(-1)

    if Args.num_buckets < 1:
        print("num_buckets must be more than 0")
        sys.exit(-1)

    if Args.every_nth < 1:
        print("every_nth must be more than 0")
        sys.exit(-1)

    if Args.excel is not None and not have_openpyxl:
        print("openpyxl is not available... use 'pip install openpyxl'")
        sys.exit(-1)

    if Args.graph and not Args.excel:
        print("can't specify -g wihout -c")
        sys.exit(-1)

# ------------------------------------------------
# Print usage.
#
def print_usage():
    print("Usage:")
    print(" -l act_storage or act_index output file")
    print("    MANDATORY - NO DEFAULT")
    print("    e.g. act_out.txt")
    print(" -h histogram to analyse")
    print("    default: depends on config read from output file")
    print(" -t analysis slice interval in seconds")
    print("    default: 3600")
    print(" -s start display from this bucket")
    print("    default: 0")
    print(" -n number of buckets to display")
    print("    default: 7")
    print(" -e show start bucket then every n-th bucket")
    print("    default: 1")
    print(" -c output Excel Spreadsheet")
    print("    NO DEFAULT")
    print(" -x (show extra information for each slice)")
    print("    default: not set")
    print(" -g add graph to Excel Spreadsheet (requires -c be set)")
    print("    default: not set")


# ------------------------------------------------
# Open log file, validate header information, and collect it for later processing.
#
def open_log_file():
    config = dict()
    misc = dict()
    act = []
    derived = []
    histogram_names = []

    config['misc'] = misc
    config['act'] = act
    config['derived'] = derived
    config['histogram_names'] = histogram_names

    # Open the log file:
    try:
        file_id = open(Args.log, "r")
    except IOError:
        print("log file " + Args.log + " not found")
        sys.exit(-1)

    misc['log_file'] = Args.log
    misc['file_id'] = file_id

    # Find and echo the version:
    line = file_id.readline()

    while line and not line.startswith("ACT version"):
        line = file_id.readline()

    if not line:
        print(Args.log + " ACT version not found")
        sys.exit(-1)

    version = line.split(" ")[2].strip()
    misc['version'] = version
    numeric_version = float(version)

    if numeric_version < 5.0 or numeric_version >= 6.0:
        print(Args.log + " ACT version not compatible")
        sys.exit(-1)

    # Find the reporting interval:
    line = file_id.readline()

    while line and not line.startswith("report-interval-sec"):
        line = file_id.readline()

    if not line:
        print("can't find report interval")
        sys.exit(-1)

    interval = long(line.split(" ")[1])

    if interval < 1:
        print("reporting interval must be more than 0")
        sys.exit(-1)

    # Find the histograms' scale:
    Hist.scale_label = " %>(ms)"
    file_id.seek(0, 0)
    line = file_id.readline()

    while line and not line.startswith("microsecond-histograms"):
        line = file_id.readline()

    if not line:
        print("can't find histograms' scale, assuming milliseconds")
        file_id.seek(0, 0)
    elif line.split(" ")[1].startswith("y"):
        Hist.scale_label = " %>(us)"

    # Adjust the slice time if necessary:
    Hist.slice_time = ((Args.slice + interval - 1) // interval) * interval
    misc['slice_time'] = Hist.slice_time


    # Echo the config from the log file:
    file_id.seek(0, 0)
    line = file_id.readline()

    while line and not line.endswith("CONFIGURATION\n"):
        line = file_id.readline()

    if not line:
        print("can't find configuration")
        sys.exit(-1)

    if line.startswith("ACT-STORAGE"):
        misc['act_type'] = "ACT-STORAGE"
        if not Args.histograms:
            Args.histograms = ["reads", "device-reads"]
    elif line.startswith("ACT-INDEX"):
        misc['act_type'] = "ACT-INDEX"
        if not Args.histograms:
            Args.histograms = ["trans-reads", "device-reads"]
    else:
        print("can't recognize configuration")
        sys.exit(-1)

    line = line.strip()

    lineno = 0
    while line:
        v = line.split(':')
        if (len(v) == 2):
            act.append([v[0], v[1].strip(), lineno])
            lineno += 1
        line = file_id.readline().strip()

    line = file_id.readline()

    while line and not line.startswith("DERIVED CONFIGURATION"):
        line = file_id.readline()

    if not line:
        print("can't find derived configuration")
        sys.exit(-1)

    line = line.strip()

    lineno = 0
    while line:
        v = line.split(':')
        if (len(v) == 2):
            derived.append([v[0], v[1].strip(), lineno])
            lineno += 1
        line = file_id.readline().strip()

    # Echo the histogram names from the log file:
    file_id.seek(0, 0)
    line = file_id.readline()

    while line and not line.startswith("HISTOGRAM NAMES\n"):
        line = file_id.readline()

    if not line:
        print("can't find histogram names")
        sys.exit(-1)

    line = line.strip()

    lineno = 0
    while line:
        if lineno > 0:
            histogram_names.append([line, lineno])
        lineno += 1
        line = file_id.readline().strip()

    return config


# ------------------------------------------------
# Print configuration information (for debugging)
#
def dump_config(config):
    for i in config:
        print ("============ dumping " + i)
        tmp = config[i]
        if i == "misc":
            for x in tmp:
                print("   %s: %s" % (x, str(tmp[x])))
        else:
            for x in tmp:
                if len(x) == 3:
                    print("   %3.3d %s: %s" % (x[-1], x[0], x[1]))
                else:
                    print("   %3.3d %s" % (x[-1], x[0]))


# ------------------------------------------------
# Print configuration information
#
def print_config(config):
    misc = config['misc']
    print("%s is ACT version %s\n" % (misc['log_file'], misc['version']))
    if Hist.slice_time != Args.slice:
        print("analyzing time slices of " + str(Hist.slice_time) + " seconds")

    print("%s CONFIGURATION" % (misc['act_type']))
    for x in config['act']:
        print("%s: %s" % (x[0], x[1]))
    print("\nDERIVED CONFIGURATION")
    for x in config['derived']:
        print("%s: %s" % (x[0], x[1]))
    print("\nHISTOGRAM NAMES")
    for x in config['histogram_names']:
        print(x[0])
    print("")
    

# ------------------------------------------------
# Output configuration information to Excel Spreadsheet
#
def excel_config(config, book, start_row, start_col):
    misc = config['misc']
    sheet = book.active

    set_cell(sheet.cell(start_row, start_col), "Version", True)
    thicken(sheet, start_row, start_col, start_row, start_col)
    set_cell(sheet.cell(start_row, start_col + 1), misc['version'], False, True)
    thicken(sheet, start_row, start_col + 1, start_row, start_col + 1)

    set_cell(sheet.cell(start_row, start_col + 3), "Log File", True)
    thicken(sheet, start_row, start_col + 3, start_row, start_col + 3)
    set_cell(sheet.cell(start_row, start_col + 4), misc['log_file'], False, True)
    thicken(sheet, start_row, start_col + 4, start_row, start_col + 5)
    merge_cells(sheet, start_row, start_col + 4, start_row, start_col + 5)
    
    today = date.today()
    set_cell(sheet.cell(start_row, start_col + 9), "ACT Report Date", True)
    thicken(sheet, start_row, start_col + 9, start_row, start_col + 10)
    merge_cells(sheet, start_row, start_col + 9, start_row, start_col + 10)

    set_cell(sheet.cell(start_row, start_col + 11), today)
    thicken(sheet, start_row, start_col + 11, start_row, start_col + 12)
    merge_cells(sheet, start_row, start_col + 11, start_row, start_col + 12)
    
    set_cell(sheet.cell(start_row, start_col + 15), "Command Line", True)
    thicken(sheet, start_row, start_col + 15, start_row, start_col + 16)
    merge_cells(sheet, start_row, start_col + 15, start_row, start_col + 16)

    set_cell(sheet.cell(start_row, start_col + 17), " ".join(sys.argv[1:]))
    thicken(sheet, start_row, start_col + 17, start_row, start_col + 21)
    merge_cells(sheet, start_row, start_col + 17, start_row, start_col + 21)

    start_row += 2

    save_row = start_row
    set_cell(sheet.cell(start_row, start_col), "HISTOGRAM NAMES", True, 'center')
    merge_cells(sheet, start_row, start_col, start_row, start_col + 1)
    start_row += 1
    for x in config['histogram_names']:
        set_cell(sheet.cell(start_row, start_col), x[0])
        merge_cells(sheet, start_row, start_col, start_row, start_col + 1)
        start_row += 1
    thicken(sheet, save_row, start_col, save_row, start_col + 1);
    thicken(sheet, save_row, start_col, start_row - 1, start_col + 1);
    max_row = start_row

    start_row = save_row
    start_col += 3
    set_cell(sheet.cell(start_row, start_col), "%s CONFIGURATION" % (misc['act_type']), True, 'center')
    merge_cells(sheet, start_row, start_col, start_row, start_col + 4)
    start_row += 1
    for x in config['derived']:
        set_cell(sheet.cell(start_row, start_col), x[0])
        merge_cells(sheet, start_row, start_col, start_row, start_col + 2)
        try:
            val = int(x[1])
        except ValueError:
            try:
                val = float(x[1])
            except ValueError:
                val = x[1]
        
        set_cell(sheet.cell(start_row, start_col + 3), val, False, True)
        merge_cells(sheet, start_row, start_col + 3, start_row, start_col + 4)
        start_row += 1
    if start_row > max_row:
        max_row = start_row
    thicken(sheet, save_row, start_col, save_row, start_col + 4);
    thicken(sheet, save_row, start_col, start_row - 1, start_col + 4);
    thicken(sheet, save_row + 1, start_col, start_row - 1, start_col + 2);

    start_row = save_row
    start_col += 6
    set_cell(sheet.cell(start_row, start_col), "DERIVED CONFIGURATION", True, 'center')
    merge_cells(sheet, start_row, start_col, start_row, start_col + 4)
    start_row += 1
    for x in config['act']:
        set_cell(sheet.cell(start_row, start_col), x[0])
        merge_cells(sheet, start_row, start_col, start_row, start_col + 2)
        try:
            val = int(x[1])
        except ValueError:
            try:
                val = float(x[1])
            except ValueError:
                val = x[1]
        
        set_cell(sheet.cell(start_row, start_col + 3), val, False, True)
        merge_cells(sheet, start_row, start_col + 3, start_row, start_col + 4)
        start_row += 1
    if start_row > max_row:
        max_row = start_row
    thicken(sheet, save_row, start_col, save_row, start_col + 4);
    thicken(sheet, save_row, start_col, start_row - 1, start_col + 4);
    thicken(sheet, save_row + 1, start_col, start_row - 1, start_col + 2);

    return max_row


# ------------------------------------------------
# Output configuration information to Excel Spreadsheet
#
def excel_graph(book, hists, start_row, start_col):
    sheet = book.active
    charts = dict()

    nc = len(Hist.display_range)
    for hist in hists:
        chart = LineChart()
        chart.title = "Request distribution for '%s'" % (hist.name)
        chart.style = 2
        chart.x_axis.title = "Time Slice"
        chart.y_axis.title = "%% exceeding 'x' %s to complete" % (Hist.scale_label[4:6])
        data = Reference(sheet, hist.start_col, Hist.first_table_row + 2,
                         hist.start_col + nc - 1, Hist.last_table_row)
        chart.add_data(data, titles_from_data=True)
        sheet.add_chart(chart, sheet.cell(Hist.last_table_row + 5, hist.start_col).coordinate)
        charts[hist.name] = chart
    
# ------------------------------------------------
# Find index + 1 of last bucket to display.
#
def find_max_bucket():
    num_buckets = Args.num_buckets

    for b in range(Args.start_bucket, ALL_BUCKETS, Args.every_nth):
        Hist.max_bucket = b + 1

        if num_buckets == 1:
            break
        else:
            num_buckets -= 1

    Hist.bucket_range = range(Hist.max_bucket)
    Hist.display_range = range(
        Args.start_bucket, Hist.max_bucket, Args.every_nth)


# ------------------------------------------------
# Print table header.
#
def print_table_header(hists):
    prefix = "slice"
    threshold_labels = ""
    threshold_underline = ""

    for i in Hist.display_range:
        threshold_labels += "%7s" % (pow(2, i))
        threshold_underline += " ------"

    if Args.extra:
        threshold_labels += "       rate"
        threshold_underline += " ----------"

    len_table = len(threshold_labels)

    for i in range(1, len(hists)):
        prev_name_len = 1 + len(hists[i - 1].name)

        if prev_name_len > len_table:
            hists[i].pre_pad = " " * (prev_name_len - len_table)

    names_out = " " * len(prefix)
    units_out = " " * len(prefix)
    labels_out = prefix
    Hist.underline = "-" * len(prefix)

    for hist in hists:
        names_out += GAP + " " + hist.name.ljust(len_table - 1)
        units_out += GAP + hist.pre_pad + Hist.scale_label.ljust(len_table)
        labels_out += GAP + hist.pre_pad + threshold_labels
        Hist.underline += GAP + hist.pre_pad + threshold_underline

    print(names_out)
    print(units_out)
    print(labels_out)
    print(Hist.underline)

def set_cell(c, v, bld=False, ralign=False):
    if (bld):
        c.font = c.font.copy(bold=True, size=12)
    else:
        c.font = c.font.copy(size=12)

    if (ralign == True):
        c.alignment = c.alignment.copy(horizontal = "right")
    elif ralign == "center":
        c.alignment = c.alignment.copy(horizontal = "center")

    c.value = v

def set_num(c, v, f, bld=False):
    c.number_format = f

    if (bld):
        c.font = c.font.copy(bold=True, size=12)
    else:
        c.font = c.font.copy(size=12)

    c.value = v

def thicken(sheet, ul_row, ul_col, lr_row, lr_col):
    for i in range(ul_col, lr_col + 1):
        c = sheet.cell(ul_row, i)
        c.border = c.border.copy(top=Side(style='medium'))

    for i in range(ul_row, lr_row + 1):
        c = sheet.cell(i, ul_col)
        c.border = c.border.copy(left=Side(style='medium'))
        
    for i in range(ul_col, lr_col + 1):
        c = sheet.cell(lr_row, i)
        c.border = c.border.copy(bottom=Side(style='medium'))

    for i in range(ul_row, lr_row + 1):
        c = sheet.cell(i, lr_col)
        c.border = c.border.copy(right=Side(style='medium'))

def merge_cells(sheet, ul_row, ul_col, lr_row, lr_col):
    left_letter = sheet.cell(ul_row, ul_col).column_letter
    right_letter = sheet.cell(ul_row, lr_col).column_letter
    fmt = "%s%d:%s%d" % (left_letter, ul_row, right_letter, lr_row)
    sheet.merge_cells(fmt)


# ------------------------------------------------
# Print table header to Excel
#
def excel_table_header(hists, book):
    slice_col = 1
    thresh_start_col = slice_col + 1
    thresh = []
    Hist.cur_row = Hist.first_table_row
    hist_len = Args.num_buckets + Args.extra + 1 # pad one column
    sheet = book.active

    for i in Hist.display_range:
        thresh.append("%d %s" % (pow(2, i), Hist.scale_label[4:6]))

    fc = Hist.first_col

    for i in range(0, len(hists)):
        tmp_col = fc + 1 + hist_len * i
        set_cell(sheet.cell(Hist.cur_row, tmp_col), hists[i].name, True)
        set_cell(sheet.cell(Hist.cur_row+1, tmp_col), Hist.scale_label, True)
        t2 = tmp_col + len(Hist.display_range) + Args.extra - 1
        merge_cells(sheet, Hist.cur_row+1, tmp_col, Hist.cur_row+1, t2)
        c = sheet.cell(Hist.cur_row+1, tmp_col)
        c.alignment = c.alignment.copy(horizontal = "center")
        thicken(sheet, Hist.cur_row+1, tmp_col, Hist.cur_row+1, t2)
        
        
    Hist.cur_row += 2

    fc = Hist.first_col
    set_cell(sheet.cell(Hist.cur_row, fc), "Slice", True, True);
    
    for i in range(0, len(hists)):
        for j in range(0, len(thresh)):
            set_cell(sheet.cell(Hist.cur_row, fc + 1 + hist_len * i + j), thresh[j], True, True)
        if (Args.extra):
            set_cell(sheet.cell(Hist.cur_row, fc + 1 + hist_len * i + len(thresh)), "Rate",
                     True, True)

    Hist.cur_row += 1

# ------------------------------------------------
# Generate latency lines.
#
def output_latency_slices(hists, file_id, book):
    # Initialization before processing time slices:
    which_slice = 0
    after_time = Hist.slice_time

    # Process all the time slices:
    while True:
        if not read_chunk(file_id, after_time, hists):
            # Note - we ignore the (possible) incomplete slice at the end.
            break

        # Print this slice's percentages over thresholds:
        which_slice += 1
        if (Args.excel is not None):
            excel_slice_line(which_slice, hists, book)
        else:
            print_slice_line(which_slice, hists)

        # Prepare for next slice:
        after_time += Hist.slice_time

    if which_slice == 0:
        print("could not find " + str(Hist.slice_time) + " seconds of data")
        sys.exit(-1)

    return which_slice


# ------------------------------------------------
# Generate latency aggregate lines.
#
def print_latency_aggregates(hists, num_slices):
    for hist in hists:
        if Args.extra:
            hist.avg_rate /= num_slices

        for i in Hist.display_range:
            hist.avg_overs[i] /= num_slices

    print(Hist.underline)
    print_avg_line(hists)
    print_max_line(hists)


# ------------------------------------------------
# Generate latency aggregate lines.
#
def excel_latency_aggregates(hists, num_slices, book):
    sheet = book.active
    cur_col = Hist.first_col
    
    # print averages
    set_cell(sheet.cell(Hist.cur_row, cur_col), "Avg", True)
    Hist.last_table_row = Hist.cur_row - 1

    for hist in hists:
        hist.start_col = cur_col + 1
        for i in Hist.display_range:
            cur_col += 1
            col_letter = sheet.cell(Hist.cur_row, cur_col).column_letter
            fmt = "=AVERAGE({0}{1}:{0}{2})".format(col_letter,
                                         Hist.first_table_row, Hist.last_table_row)
            set_num(sheet.cell(Hist.cur_row, cur_col), fmt, "0.00%", True)

        if (Args.extra):
            cur_col += 1
            col_letter = sheet.cell(Hist.cur_row, cur_col).column_letter
            fmt = "=AVERAGE({0}{1}:{0}{2})".format(col_letter,
                                                   Hist.first_table_row,
                                                   Hist.last_table_row)
            set_num(sheet.cell(Hist.cur_row, cur_col), fmt, "0.0", True)

        cur_col += 1

    Hist.cur_row += 1

    # print max
    cur_col = Hist.first_col
    set_cell(sheet.cell(Hist.cur_row, cur_col), "Max", True)

    for hist in hists:
        for i in Hist.display_range:
            cur_col += 1
            col_letter = sheet.cell(Hist.cur_row, cur_col).column_letter
            fmt = "=MAX({0}{1}:{0}{2})".format(col_letter,
                                         Hist.first_table_row, Hist.last_table_row)
            set_num(sheet.cell(Hist.cur_row, cur_col), fmt, "0.00%", True)

        if (Args.extra):
            cur_col += 1
            col_letter = sheet.cell(Hist.cur_row, cur_col).column_letter
            fmt = "=MAX({0}{1}:{0}{2})".format(col_letter,
                                         Hist.first_table_row, Hist.last_table_row)
            set_num(sheet.cell(Hist.cur_row, cur_col), fmt, "0.0", True)

        cur_col += 1

            
# ------------------------------------------------
# Get the data chunk reported by act at the specified after_time.
#
def read_chunk(file_id, after_time, hists):
    find_line = "after " + str(after_time) + " "

    while True:
        line = file_id.readline()

        if not line:
            return False

        if line.startswith(find_line):
            break

    got_chunk = False
    line = file_id.readline()

    while line and line.strip():
        for hist in hists:
            if line.startswith(hist.name):
                line = read_bucket_values(line, file_id, hist)
                got_chunk = True
                break
        else:
            line = file_id.readline()

    return got_chunk


# ------------------------------------------------
# Print a latency data output line.
#
def print_slice_line(slice_tag, hists):
    output = "%5s" % (slice_tag)

    for hist in hists:
        output += GAP + hist.pre_pad

        for i in Hist.display_range:
            output += "%7.2f" % (hist.overs[i])

        if Args.extra:
            output += "%11.1f" % (hist.rate)

    print(output)


# ------------------------------------------------
# Print a latency data output line.
#
def excel_slice_line(slice_tag, hists, book):
    sheet = book.active
    cur_col = Hist.first_col
    set_num(sheet.cell(Hist.cur_row, cur_col), slice_tag, "00", True)

    for hist in hists:
        for i in Hist.display_range:
            cur_col += 1
            set_num(sheet.cell(Hist.cur_row, cur_col), hist.overs[i] / 100.0, "0.00%")

        if (Args.extra):
            cur_col += 1
            set_num(sheet.cell(Hist.cur_row, cur_col), hist.rate, "0.0")

        cur_col += 1

    Hist.cur_row += 1



# ------------------------------------------------
# Print a latency average data output line.
#
def print_avg_line(hists):
    output = "  avg"

    for hist in hists:
        output += GAP + hist.pre_pad

        for i in Hist.display_range:
            output += "%7.2f" % (hist.avg_overs[i])

        if Args.extra:
            output += "%11.1f" % (hist.avg_rate)

    print(output)


# ------------------------------------------------
# Print a latency maximum data output line.
#
def print_max_line(hists):
    output = "  max"

    for hist in hists:
        output += GAP + hist.pre_pad

        for i in Hist.display_range:
            output += "%7.2f" % (hist.max_overs[i])

        if Args.extra:
            output += "%11.1f" % (hist.max_rate)

    print(output)


# ------------------------------------------------
# Get one set of bucket values.
#
def read_bucket_values(line, file_id, hist):
    values = [0] * Hist.max_bucket
    total, line = read_total_ops(line, file_id)
    b_min = 0

    while True:
        found = 0

        for b in Hist.bucket_range[b_min:]:
            r = BUCKET_PATTERNS[b]

            if r.search(line):
                found += 1
                values[b] = long(r.search(line).group(1))

        if found == 0:
            break

        line = file_id.readline()
        b_min += found

    hist.slice_total = total - hist.old_total
    hist.slice_values = [a - b for a, b in zip(values, hist.old_values)]
    hist.old_total = total
    hist.old_values = values
    bucket_percentages_over(hist)
    bucket_aggregations(hist)

    return line


# ------------------------------------------------
# Parse a histogram total from a act output line.
#
def read_total_ops(line, file_id):
    total = long(line[line.find("(") + 1: line.find(" total)")])
    line = file_id.readline()

    return total, line


# ------------------------------------------------
# Get the percentage excesses for every bucket.
#
def bucket_percentages_over(hist):
    hist.overs = [0.0] * Hist.max_bucket

    if hist.slice_total == 0:
        return

    delta = 0

    for b in Hist.bucket_range:
        delta += hist.slice_values[b]
        hist.overs[b] = round(
            ((hist.slice_total - delta) * 100.0) / hist.slice_total, 2)


# ------------------------------------------------
# Track maximums and totals to calculate averages.
#
def bucket_aggregations(hist):
    hist.rate = round(float(hist.slice_total) / Hist.slice_time, 1)

    if Args.extra:
        hist.avg_rate += hist.rate

        if hist.rate > hist.max_rate:
            hist.max_rate = hist.rate

    for i in Hist.display_range:
        hist.avg_overs[i] += hist.overs[i]

        if hist.overs[i] > hist.max_overs[i]:
            hist.max_overs[i] = hist.overs[i]


# ==========================================================
# Execution.
#

if __name__ == "__main__":
    main()
